"""
Excel file reading functionality.

Extracts a rectangular region from an Excel file and returns it as a pandas DataFrame.
"""

import re
from typing import Tuple, Optional
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import MergedCell


class ExcelReaderError(Exception):
    """Base exception for Excel reading errors."""

    pass


class InvalidRangeError(ExcelReaderError):
    """Raised when the cell range specification is invalid."""

    pass


class SheetNotFoundError(ExcelReaderError):
    """Raised when the specified sheet doesn't exist."""

    pass


class FileReadError(ExcelReaderError):
    """Raised when the Excel file cannot be read."""

    pass


def parse_cell_reference(cell_ref: str) -> Tuple[int, int]:
    """
    Parse a cell reference like 'B2' into (row, column) tuple.

    Args:
        cell_ref: Excel cell reference (e.g., 'B2', 'AA100')

    Returns:
        Tuple of (row_number, column_number), both 1-indexed

    Raises:
        InvalidRangeError: If the cell reference is invalid
    """
    # Match pattern: one or more letters followed by one or more digits
    match = re.match(r"^([A-Za-z]+)(\d+)$", cell_ref.strip())
    if not match:
        raise InvalidRangeError(f"Invalid cell reference: '{cell_ref}'")

    col_letters = match.group(1).upper()
    row_num = int(match.group(2))

    if row_num < 1:
        raise InvalidRangeError(f"Row number must be positive: '{cell_ref}'")

    # Convert column letters to number (A=1, B=2, ..., Z=26, AA=27, etc.)
    col_num = 0
    for char in col_letters:
        col_num = col_num * 26 + (ord(char) - ord("A") + 1)

    return row_num, col_num


def normalize_range(
    start_cell: str, end_cell: str
) -> Tuple[Tuple[int, int], Tuple[int, int]]:
    """
    Normalize a cell range, ensuring start is top-left and end is bottom-right.

    Args:
        start_cell: Start cell reference (e.g., 'B2')
        end_cell: End cell reference (e.g., 'F45')

    Returns:
        Tuple of ((start_row, start_col), (end_row, end_col))
    """
    start_row, start_col = parse_cell_reference(start_cell)
    end_row, end_col = parse_cell_reference(end_cell)

    # Normalize so start is always top-left
    min_row = min(start_row, end_row)
    max_row = max(start_row, end_row)
    min_col = min(start_col, end_col)
    max_col = max(start_col, end_col)

    return (min_row, min_col), (max_row, max_col)


def get_merged_cell_value(ws: Worksheet, row: int, col: int) -> Optional[any]:
    """
    Get the value for a cell, handling merged cells.

    For merged cells, returns the value from the top-left cell of the merged range.

    Args:
        ws: openpyxl Worksheet
        row: Row number (1-indexed)
        col: Column number (1-indexed)

    Returns:
        The cell value, or None if empty
    """
    cell = ws.cell(row=row, column=col)

    # If this is a merged cell, find the master cell
    if isinstance(cell, MergedCell):
        # Find which merged range this cell belongs to
        for merged_range in ws.merged_cells.ranges:
            if (row, col) in [
                (r, c)
                for r in range(merged_range.min_row, merged_range.max_row + 1)
                for c in range(merged_range.min_col, merged_range.max_col + 1)
            ]:
                # Return the value from the top-left cell
                return ws.cell(
                    row=merged_range.min_row, column=merged_range.min_col
                ).value

    return cell.value


def read_header_metadata(
    ws: Worksheet,
    header_row: int,
    start_col: int,
    end_col: int,
) -> dict[int, str]:
    """
    Read the row above the headers to extract column metadata.

    Handles merged cells - if a merged cell spans multiple columns,
    all those columns get the same metadata value.

    Args:
        ws: openpyxl Worksheet
        header_row: The row number of the headers (1-indexed)
        start_col: Starting column (1-indexed)
        end_col: Ending column (1-indexed)

    Returns:
        Dictionary mapping column index (0-based) to metadata string
    """
    metadata_row = header_row - 1
    if metadata_row < 1:
        return {}  # No row above headers

    metadata = {}
    col_index = 0

    for col_idx in range(start_col, end_col + 1):
        value = get_merged_cell_value(ws, metadata_row, col_idx)
        if value is not None and str(value).strip():
            metadata[col_index] = str(value).strip()
        col_index += 1

    return metadata


def read_rectangle(
    file_path: str,
    sheet_name: str,
    start_cell: str,
    end_cell: str,
    fill_merged: bool = True,
    include_header_metadata: bool = True,
) -> tuple[pd.DataFrame, dict[int, str]]:
    """
    Read a rectangular region from an Excel file.

    The first row of the rectangle becomes the column headers.
    Remaining rows become the data.

    Args:
        file_path: Path to the Excel file
        sheet_name: Name of the sheet to read from
        start_cell: Top-left cell of the rectangle (e.g., 'B2')
        end_cell: Bottom-right cell of the rectangle (e.g., 'F45')
        fill_merged: If True, fill merged cells with the merged value.
                    If False, only the top-left cell of a merge has a value.
        include_header_metadata: If True, also read one row above headers
                                for column metadata (descriptions, categories).

    Returns:
        Tuple of (DataFrame, header_metadata_dict)
        - DataFrame: pandas DataFrame with the extracted data
        - header_metadata_dict: dict mapping column index (0-based) to metadata string
                               from the row above headers (empty dict if none found)

    Raises:
        FileReadError: If the file cannot be read
        SheetNotFoundError: If the sheet doesn't exist
        InvalidRangeError: If the range is invalid
    """
    # Validate file exists
    path = Path(file_path)
    if not path.exists():
        raise FileReadError(f"File not found: {file_path}")

    if not path.suffix.lower() in [".xlsx", ".xlsm", ".xltx", ".xltm"]:
        raise FileReadError(
            f"Unsupported file format: {path.suffix}. Use .xlsx or similar."
        )

    # Parse and normalize the range
    (start_row, start_col), (end_row, end_col) = normalize_range(start_cell, end_cell)

    # Load workbook
    try:
        # data_only=True to get computed values instead of formulas
        wb = load_workbook(file_path, data_only=True, read_only=False)
    except Exception as e:
        raise FileReadError(f"Cannot read Excel file: {e}")

    # Get the sheet
    if sheet_name not in wb.sheetnames:
        available = ", ".join(wb.sheetnames)
        raise SheetNotFoundError(
            f"Sheet '{sheet_name}' not found. Available sheets: {available}"
        )

    ws = wb[sheet_name]

    # Read header metadata (row above headers) if requested
    header_metadata = {}
    if include_header_metadata:
        header_metadata = read_header_metadata(ws, start_row, start_col, end_col)

    # Read the data
    data = []
    for row_idx in range(start_row, end_row + 1):
        row_data = []
        for col_idx in range(start_col, end_col + 1):
            if fill_merged:
                value = get_merged_cell_value(ws, row_idx, col_idx)
            else:
                cell = ws.cell(row=row_idx, column=col_idx)
                value = cell.value if not isinstance(cell, MergedCell) else None
            row_data.append(value)
        data.append(row_data)

    wb.close()

    # Handle empty data
    if not data:
        raise InvalidRangeError(f"No data found in range {start_cell}:{end_cell}")

    # First row is headers
    headers = data[0]

    # Handle duplicate and empty column names
    final_headers = []
    seen = {}
    for i, h in enumerate(headers):
        # Convert to string, handle None/empty
        if h is None or (isinstance(h, str) and h.strip() == ""):
            h = f"column_{i + 1}"
        else:
            h = str(h)

        # Handle duplicates
        original = h
        count = seen.get(h, 0)
        if count > 0:
            h = f"{original}_{count + 1}"
        seen[original] = count + 1
        final_headers.append(h)

    # Remaining rows are data
    if len(data) == 1:
        # Headers only, no data rows
        df = pd.DataFrame(columns=final_headers)
    else:
        df = pd.DataFrame(data[1:], columns=final_headers)

    return df, header_metadata


def list_sheets(file_path: str) -> list[str]:
    """
    List all sheet names in an Excel file.

    Args:
        file_path: Path to the Excel file

    Returns:
        List of sheet names
    """
    path = Path(file_path)
    if not path.exists():
        raise FileReadError(f"File not found: {file_path}")

    try:
        wb = load_workbook(file_path, read_only=True)
        sheets = wb.sheetnames
        wb.close()
        return sheets
    except Exception as e:
        raise FileReadError(f"Cannot read Excel file: {e}")


def detect_data_range(
    file_path: str, sheet_name: str, start_cell: str = "A1"
) -> Tuple[str, str]:
    """
    Detect the extent of data starting from a given cell.

    Useful when you don't know the exact end of the data.

    Args:
        file_path: Path to the Excel file
        sheet_name: Name of the sheet
        start_cell: Cell to start detection from

    Returns:
        Tuple of (start_cell, end_cell) representing the data range
    """
    path = Path(file_path)
    if not path.exists():
        raise FileReadError(f"File not found: {file_path}")

    try:
        wb = load_workbook(file_path, data_only=True)
    except Exception as e:
        raise FileReadError(f"Cannot read Excel file: {e}")

    if sheet_name not in wb.sheetnames:
        available = ", ".join(wb.sheetnames)
        raise SheetNotFoundError(
            f"Sheet '{sheet_name}' not found. Available sheets: {available}"
        )

    ws = wb[sheet_name]

    start_row, start_col = parse_cell_reference(start_cell)

    # Find the last row with data
    max_row = start_row
    for row in range(start_row, ws.max_row + 1):
        has_data = False
        for col in range(start_col, ws.max_column + 1):
            if ws.cell(row=row, column=col).value is not None:
                has_data = True
                break
        if has_data:
            max_row = row
        else:
            # Allow a few empty rows before stopping
            empty_count = 0
            for check_row in range(row, min(row + 5, ws.max_row + 1)):
                row_empty = True
                for col in range(start_col, ws.max_column + 1):
                    if ws.cell(row=check_row, column=col).value is not None:
                        row_empty = False
                        break
                if row_empty:
                    empty_count += 1
                else:
                    break
            if empty_count >= 3:
                break
            max_row = row

    # Find the last column with data
    max_col = start_col
    for col in range(start_col, ws.max_column + 1):
        has_data = False
        for row in range(start_row, max_row + 1):
            if ws.cell(row=row, column=col).value is not None:
                has_data = True
                break
        if has_data:
            max_col = col

    wb.close()

    # Convert back to cell references
    def col_to_letter(col_num: int) -> str:
        result = ""
        while col_num > 0:
            col_num, remainder = divmod(col_num - 1, 26)
            result = chr(65 + remainder) + result
        return result

    end_cell = f"{col_to_letter(max_col)}{max_row}"

    return start_cell, end_cell
