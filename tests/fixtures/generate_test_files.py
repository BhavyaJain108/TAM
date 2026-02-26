"""
Generate test Excel files for the ingestion pipeline.

Run this script to create all test fixtures:
    python generate_test_files.py

Requires: pip install openpyxl pandas
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment
from datetime import datetime, date
import os

OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))


def create_clean_test():
    """test_clean.xlsx - Happy path with clean, well-formatted data."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Clients"

    # Clean data starting at B2
    data = [
        ["Client Name", "Industry", "Region", "Annual Revenue", "Status", "Start Date"],
        ["Acme Corporation", "Technology", "North America", 2500000, "Active", "2020-01-15"],
        ["Beta Industries", "Manufacturing", "Europe", 1800000, "Active", "2019-06-01"],
        ["Gamma Health", "Healthcare", "North America", 3200000, "Active", "2021-03-10"],
        ["Delta Financial", "Finance", "Asia Pacific", 4500000, "Active", "2018-11-20"],
        ["Epsilon Retail", "Retail", "Europe", 900000, "Prospect", "2023-02-28"],
        ["Zeta Logistics", "Transportation", "North America", 1500000, "Active", "2020-08-15"],
        ["Eta Media", "Media", "Europe", 2100000, "Dormant", "2017-04-01"],
        ["Theta Energy", "Energy", "Asia Pacific", 5800000, "Active", "2019-09-12"],
        ["Iota Foods", "Consumer Goods", "North America", 750000, "Prospect", "2024-01-05"],
        ["Kappa Software", "Technology", "Europe", 3400000, "Active", "2022-07-18"],
    ]

    # Write starting at B2
    for row_idx, row in enumerate(data, start=2):
        for col_idx, value in enumerate(row, start=2):
            ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(os.path.join(OUTPUT_DIR, "test_clean.xlsx"))
    print("Created: test_clean.xlsx")


def create_messy_test():
    """test_messy.xlsx - All the common data quality issues."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Deals"

    # Messy data with various issues
    data = [
        ["Acct", "Opp Name", "$ Val", "Stage", "Exp Close", "Lead"],
        ["Acme", "Proj Alpha", "2.5M", "CW", "2024-03-15", "J. Smith"],
        ["  Acme  ", "Proj Beta", "900K", "Neg", "03/15/2024", "J. Smith"],  # whitespace, different date format
        ["BETA HEALTH", "Digital Tx", "$1.8M", "Prop", "2024-09-01", "A. Jones"],  # different casing, $ symbol
        ["???", "Proj Gamma", "500K", "Lead", "???", "M. Chen"],  # placeholders
        ["beta health", "Expansion", "1,200,000", "CW", "Sep 1, 2024", "A. Jones"],  # lowercase, comma format, text date
        ["Gamma", "New Initiative", "N/A", "Lead", "TBD", " R. Lee"],  # N/A, TBD, leading space
        ["", "Mystery Deal", "750K", "Neg", "2024-06-01", ""],  # empty values
        ["Delta Co", "Project X", "2.5M", "Prop", "2024-12-31", "J. smith"],  # lowercase surname
        ["Acme", "Renewal", "(500K)", "CW", "2024-01-01", "J. Smith"],  # negative in parens
        ["GAMMA", "Phase 2", "3M", "Neg", None, "R. Lee"],  # null value
    ]

    for row_idx, row in enumerate(data, start=2):
        for col_idx, value in enumerate(row, start=2):
            ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(os.path.join(OUTPUT_DIR, "test_messy.xlsx"))
    print("Created: test_messy.xlsx")


def create_merged_test():
    """test_merged.xlsx - Merged cells in headers and data."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    # Title row (merged)
    ws.merge_cells('B1:G1')
    ws['B1'] = "Q1 2024 Sales Report"
    ws['B1'].font = Font(bold=True, size=14)
    ws['B1'].alignment = Alignment(horizontal='center')

    # Merged header group
    ws.merge_cells('B2:C2')
    ws['B2'] = "Client Info"
    ws.merge_cells('D2:E2')
    ws['D2'] = "Deal Info"
    ws.merge_cells('F2:G2')
    ws['F2'] = "Financial"

    # Actual column headers
    headers = ["Client", "Region", "Deal", "Stage", "Q1 Revenue", "Q1 Margin"]
    for col_idx, header in enumerate(headers, start=2):
        ws.cell(row=3, column=col_idx, value=header)

    # Data with a merged cell in the data area
    data = [
        ["Acme Corp", "NA", "Project A", "Won", 500000, 0.25],
        ["Acme Corp", "NA", "Project B", "Won", 300000, 0.30],
        ["Beta Inc", "EU", "Expansion", "Lost", 0, 0],
        ["Gamma LLC", "APAC", "New Deal", "Active", 250000, 0.22],
    ]

    for row_idx, row in enumerate(data, start=4):
        for col_idx, value in enumerate(row, start=2):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Merge a data cell (Acme Corp appears twice - merge it)
    ws.merge_cells('B4:B5')
    ws['B4'].alignment = Alignment(vertical='center')
    ws.merge_cells('C4:C5')  # Also merge region

    wb.save(os.path.join(OUTPUT_DIR, "test_merged.xlsx"))
    print("Created: test_merged.xlsx")


def create_pivot_test():
    """test_pivot.xlsx - Pivot table layout with time periods as columns."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Revenue by Region"

    data = [
        ["Region", "Q1 2023", "Q2 2023", "Q3 2023", "Q4 2023", "Q1 2024"],
        ["North America", 1200000, 1350000, 1400000, 1550000, 1600000],
        ["Europe", 800000, 850000, 900000, 950000, 1000000],
        ["Asia Pacific", 600000, 750000, 900000, 1100000, 1300000],
        ["Latin America", 200000, 220000, 240000, 260000, 280000],
        ["Middle East", 150000, 175000, 200000, 225000, 250000],
    ]

    for row_idx, row in enumerate(data, start=2):
        for col_idx, value in enumerate(row, start=2):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Add a totals row
    ws.cell(row=8, column=2, value="Total")
    for col in range(3, 8):
        ws.cell(row=8, column=col, value=f"=SUM({chr(64+col)}3:{chr(64+col)}7)")

    wb.save(os.path.join(OUTPUT_DIR, "test_pivot.xlsx"))
    print("Created: test_pivot.xlsx")


def create_formulas_test():
    """test_formulas.xlsx - Cells with various formulas."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Calculations"

    # Headers
    headers = ["Item", "Quantity", "Unit Price", "Total", "Tax", "Grand Total"]
    for col_idx, header in enumerate(headers, start=2):
        ws.cell(row=2, column=col_idx, value=header)

    # Data with formulas
    items = [
        ("Widget A", 100, 25.00),
        ("Widget B", 50, 45.00),
        ("Widget C", 200, 12.50),
        ("Service Fee", 1, 500.00),
    ]

    for row_idx, (item, qty, price) in enumerate(items, start=3):
        ws.cell(row=row_idx, column=2, value=item)
        ws.cell(row=row_idx, column=3, value=qty)
        ws.cell(row=row_idx, column=4, value=price)
        ws.cell(row=row_idx, column=5, value=f"=C{row_idx}*D{row_idx}")  # Total
        ws.cell(row=row_idx, column=6, value=f"=E{row_idx}*0.08")  # Tax
        ws.cell(row=row_idx, column=7, value=f"=E{row_idx}+F{row_idx}")  # Grand Total

    # Summary row
    ws.cell(row=8, column=2, value="TOTAL")
    ws.cell(row=8, column=5, value="=SUM(E3:E6)")
    ws.cell(row=8, column=6, value="=SUM(F3:F6)")
    ws.cell(row=8, column=7, value="=SUM(G3:G6)")

    wb.save(os.path.join(OUTPUT_DIR, "test_formulas.xlsx"))
    print("Created: test_formulas.xlsx")


def create_types_test():
    """test_types.xlsx - One column per data type edge case."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Types"

    # This tests type detection edge cases
    headers = [
        "clean_int", "clean_float", "int_as_float", "mixed_num_str",
        "date_iso", "date_mixed", "bool_native", "bool_string",
        "currency", "percentage", "scientific", "all_null",
        "single_value", "with_placeholders"
    ]

    for col_idx, header in enumerate(headers, start=2):
        ws.cell(row=2, column=col_idx, value=header)

    # Row 1 data
    row3 = [1, 1.5, 1.0, 1, "2024-01-15", "2024-01-15", True, "Yes", "$1,000", "50%", "1.5e6", None, "Active", "Value1"]
    # Row 2 data
    row4 = [2, 2.7, 2.0, 2, "2024-02-20", "02/20/2024", False, "No", "$2,500", "75%", "2.3e-4", None, "Active", "???"]
    # Row 3 data
    row5 = [3, 3.14, 3.0, "three", "2024-03-25", "March 25, 2024", True, "YES", "€3,000", "100%", "9.8e2", None, "Active", "N/A"]
    # Row 4 data
    row6 = [4, 4.0, 4.0, 4, "2024-04-30", "2024-04-30", False, "no", "$4,500", "25.5%", "1e10", None, "Active", "TBD"]
    # Row 5 data
    row7 = [5, 5.5, 5.0, "5", "2024-05-15", "May 15 2024", True, "Y", "$5,000", "0%", "3.14e0", None, "Active", "Value5"]

    for row_idx, row_data in enumerate([row3, row4, row5, row6, row7], start=3):
        for col_idx, value in enumerate(row_data, start=2):
            ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(os.path.join(OUTPUT_DIR, "test_types.xlsx"))
    print("Created: test_types.xlsx")


def create_unicode_test():
    """test_unicode.xlsx - International characters and emoji."""
    wb = Workbook()
    ws = wb.active
    ws.title = "International"

    data = [
        ["客户名称", "Kundenname", "العميل", "Status", "Notes"],
        ["上海科技", "München GmbH", "شركة الرياض", "Active ✓", "Primary contact: 田中太郎"],
        ["北京贸易", "Zürich AG", "دبي للتجارة", "Pending ⏳", "Follow up: próxima semana"],
        ["东京电子", "Société Paris", "القاهرة تك", "Inactive ✗", "Last contact: été 2023"],
        ["서울테크", "Øresund Corp", "تونس ديجيتال", "Active ✓", "VIP customer 🌟"],
    ]

    for row_idx, row in enumerate(data, start=2):
        for col_idx, value in enumerate(row, start=2):
            ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(os.path.join(OUTPUT_DIR, "test_unicode.xlsx"))
    print("Created: test_unicode.xlsx")


def create_multi_table_test():
    """test_multi_table.xlsx - Multiple related tables across sheets."""
    wb = Workbook()

    # Sheet 1: Clients
    ws1 = wb.active
    ws1.title = "Clients"
    clients = [
        ["Client ID", "Client Name", "Industry", "Region"],
        ["C001", "Acme Corporation", "Technology", "North America"],
        ["C002", "Beta Industries", "Manufacturing", "Europe"],
        ["C003", "Gamma Health", "Healthcare", "Asia Pacific"],
        ["C004", "Delta Financial", "Finance", "North America"],
    ]
    for row_idx, row in enumerate(clients, start=2):
        for col_idx, value in enumerate(row, start=2):
            ws1.cell(row=row_idx, column=col_idx, value=value)

    # Sheet 2: Deals (references Clients)
    ws2 = wb.create_sheet("Deals")
    deals = [
        ["Deal ID", "Client ID", "Deal Name", "Value", "Stage"],
        ["D001", "C001", "Project Alpha", 500000, "Won"],
        ["D002", "C001", "Project Beta", 300000, "Active"],
        ["D003", "C002", "Expansion", 750000, "Lost"],
        ["D004", "C003", "Digital Transform", 1200000, "Active"],
        ["D005", "C004", "Advisory", 200000, "Won"],
    ]
    for row_idx, row in enumerate(deals, start=2):
        for col_idx, value in enumerate(row, start=2):
            ws2.cell(row=row_idx, column=col_idx, value=value)

    # Sheet 3: Contacts (references Clients)
    ws3 = wb.create_sheet("Contacts")
    contacts = [
        ["Contact ID", "Client ID", "Name", "Title", "Email"],
        ["CT001", "C001", "John Smith", "CEO", "john@acme.com"],
        ["CT002", "C001", "Jane Doe", "CFO", "jane@acme.com"],
        ["CT003", "C002", "Hans Mueller", "Director", "hans@beta.de"],
        ["CT004", "C003", "Sarah Chen", "VP Sales", "sarah@gamma.com"],
        ["CT005", "C004", "Mike Johnson", "Partner", "mike@delta.com"],
    ]
    for row_idx, row in enumerate(contacts, start=2):
        for col_idx, value in enumerate(row, start=2):
            ws3.cell(row=row_idx, column=col_idx, value=value)

    wb.save(os.path.join(OUTPUT_DIR, "test_multi_table.xlsx"))
    print("Created: test_multi_table.xlsx")


def create_edge_cases_test():
    """test_edge_cases.xlsx - Boundary conditions and unusual scenarios."""
    wb = Workbook()

    # Sheet 1: Single row (headers only effectively)
    ws1 = wb.active
    ws1.title = "SingleRow"
    ws1['B2'] = "Col A"
    ws1['C2'] = "Col B"
    ws1['D2'] = "Col C"
    ws1['B3'] = "Only"
    ws1['C3'] = "One"
    ws1['D3'] = "Row"

    # Sheet 2: Single column
    ws2 = wb.create_sheet("SingleColumn")
    ws2['B2'] = "Values"
    for i, val in enumerate(["Alpha", "Beta", "Gamma", "Delta"], start=3):
        ws2.cell(row=i, column=2, value=val)

    # Sheet 3: All same value
    ws3 = wb.create_sheet("Constant")
    ws3['B2'] = "Status"
    ws3['C2'] = "Code"
    for i in range(3, 13):
        ws3.cell(row=i, column=2, value="Active")
        ws3.cell(row=i, column=3, value="A")

    # Sheet 4: High cardinality (exactly 20 unique - boundary)
    ws4 = wb.create_sheet("Boundary20")
    ws4['B2'] = "ID"
    ws4['C2'] = "Value"
    for i in range(1, 21):
        ws4.cell(row=i+2, column=2, value=f"ID_{i:03d}")
        ws4.cell(row=i+2, column=3, value=f"Value_{i}")

    # Sheet 5: 21 unique (just over boundary)
    ws5 = wb.create_sheet("Boundary21")
    ws5['B2'] = "ID"
    ws5['C2'] = "Value"
    for i in range(1, 22):
        ws5.cell(row=i+2, column=2, value=f"ID_{i:03d}")
        ws5.cell(row=i+2, column=3, value=f"Value_{i}")

    # Sheet 6: Column name edge cases
    ws6 = wb.create_sheet("ColumnNames")
    problem_headers = [
        "Normal",
        "With Spaces",
        "$ Special & Chars!",
        "Client/Industry",
        "2024 Revenue",  # starts with number
        "SELECT",  # SQL keyword
        "from",  # SQL keyword lowercase
        "",  # empty
        "Name",  # will be duplicate
        "Name",  # duplicate
    ]
    for col_idx, header in enumerate(problem_headers, start=2):
        ws6.cell(row=2, column=col_idx, value=header if header else None)
        ws6.cell(row=3, column=col_idx, value=f"val_{col_idx}")
        ws6.cell(row=4, column=col_idx, value=f"val_{col_idx}_2")

    wb.save(os.path.join(OUTPUT_DIR, "test_edge_cases.xlsx"))
    print("Created: test_edge_cases.xlsx")


def create_large_test():
    """test_large.xlsx - Performance testing with many rows."""
    wb = Workbook()
    ws = wb.active
    ws.title = "LargeData"

    # Headers
    headers = ["ID", "Name", "Category", "Value", "Date", "Status"]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    # Generate 10,000 rows (for quick tests; increase for stress testing)
    categories = ["A", "B", "C", "D", "E"]
    statuses = ["Active", "Pending", "Closed", "On Hold"]

    import random
    random.seed(42)

    for i in range(1, 10001):
        ws.cell(row=i+1, column=1, value=i)
        ws.cell(row=i+1, column=2, value=f"Entity_{i:05d}")
        ws.cell(row=i+1, column=3, value=random.choice(categories))
        ws.cell(row=i+1, column=4, value=round(random.uniform(100, 10000), 2))
        ws.cell(row=i+1, column=5, value=f"2024-{random.randint(1,12):02d}-{random.randint(1,28):02d}")
        ws.cell(row=i+1, column=6, value=random.choice(statuses))

    wb.save(os.path.join(OUTPUT_DIR, "test_large.xlsx"))
    print("Created: test_large.xlsx (10,000 rows)")


if __name__ == "__main__":
    print("Generating test Excel files...\n")

    create_clean_test()
    create_messy_test()
    create_merged_test()
    create_pivot_test()
    create_formulas_test()
    create_types_test()
    create_unicode_test()
    create_multi_table_test()
    create_edge_cases_test()
    create_large_test()

    print("\nAll test files created successfully!")
    print(f"Location: {OUTPUT_DIR}")
